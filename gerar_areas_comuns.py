# -*- coding: utf-8 -*-
"""
Prepara os dados do relatorio de AREAS COMUNS.

A aba "AREAS COMUNS" e um historico largo: cada linha e um ponto de medicao
(guarita, quadra, ADM, ...) e cada coluna e uma data de leitura (~47 datas,
2023->2026). O arquivo mais recente contem TODAS as datas (superset), entao
lemos apenas essa aba.

Saida: relatorio/dados_areas_comuns.js  ->  window.DADOS_AC = {...}

Uso:
    python gerar_areas_comuns.py
"""

import glob
import json
import os
import re
import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SAIDA = os.path.join(BASE_DIR, "relatorio", "dados_areas_comuns.js")

import openpyxl

GRUPOS = ["MATA", "1ª ENCOSTA", "2ª ENCOSTA", "3ª ENCOSTA", "CAMPO", "COLINA", "DEMAIS LOCALIDADES"]
GRUPOS_UP = {g.upper() for g in GRUPOS}
# Ponto que nao pertence as areas comuns (codigo de unidade residencial, medidor removido).
RE_UNIDADE = re.compile(r"^[A-Z]\d-\d{3}$")
MES3 = ["", "jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]


def rotulo(dt):
    return f"{dt.day:02d}/{MES3[dt.month]}/{str(dt.year)[2:]}"


def num(v):
    """Leitura -> numero ou None (textos como '12 L' viram None)."""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return v
    return None


def escolher_arquivo():
    """Escolhe o arquivo cuja aba AREAS COMUNS tem mais datas (superset)."""
    melhor, melhor_n = None, -1
    for f in glob.glob(os.path.join(BASE_DIR, "**", "*PARA COBRAN*.xlsx"), recursive=True):
        if os.path.basename(f).startswith("~$"):
            continue
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        except Exception:
            continue
        if "ÁREAS COMUNS" in wb.sheetnames:
            ws = wb["ÁREAS COMUNS"]
            n = sum(1 for c in range(1, ws.max_column + 1)
                    if isinstance(ws.cell(2, c).value, (datetime.datetime, datetime.date)))
            if n > melhor_n:
                melhor, melhor_n = f, n
        wb.close()
    return melhor


def main():
    arq = escolher_arquivo()
    if not arq:
        raise SystemExit("Nenhuma aba 'ÁREAS COMUNS' encontrada.")
    wb = openpyxl.load_workbook(arq, data_only=True)
    ws = wb["ÁREAS COMUNS"]
    H = 2

    # localizar coluna LOCAL e as colunas de data
    c_local = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(H, c).value
        if isinstance(v, str) and v.strip().upper() == "LOCAL":
            c_local = c
            break
    if c_local is None:
        c_local = 2
    cols_data = [(c, ws.cell(H, c).value) for c in range(c_local + 1, ws.max_column + 1)
                 if isinstance(ws.cell(H, c).value, (datetime.datetime, datetime.date))]
    datas = [d.date() if isinstance(d, datetime.datetime) else d for _, d in cols_data]
    col_nums = [c for c, _ in cols_data]

    TARIFA = 4.76

    # ---- percorre as linhas ----
    areas = []            # cada ponto de medicao valido
    grupo_atual = "Outros"
    total_por_col = [0.0] * len(col_nums)   # consumo agregado atribuido a cada coluna de data
    excluidos = []

    for r in range(H + 1, ws.max_row + 1):
        loc = ws.cell(r, c_local).value
        if loc is None or not str(loc).strip():
            continue
        loc = str(loc).strip()
        reads = [num(ws.cell(r, c).value) for c in col_nums]
        tem_num = any(v is not None for v in reads)

        if loc.upper() in GRUPOS_UP and not tem_num:
            grupo_atual = loc
            continue
        if not tem_num:
            continue  # linha de nota / cabecalho sem leituras
        if RE_UNIDADE.match(loc):
            excluidos.append(loc)  # ex: B0-080 (medidor removido, nao e area comum)
            continue

        # ---- consumo alinhado as colunas de data: consumo[i] = leitura[i] - leitura[i-1]
        #      (so quando ambas as colunas tem leitura numerica) ----
        col_consumo = [None] * len(reads)
        for i in range(1, len(reads)):
            if reads[i] is not None and reads[i - 1] is not None:
                col_consumo[i] = reads[i] - reads[i - 1]
                total_por_col[i] += col_consumo[i]

        serie = []
        for i, v in enumerate(reads):
            if v is None:
                continue
            consumo = col_consumo[i]
            flags = []
            if consumo is not None and consumo < 0:
                flags.append("consumo_negativo")
            serie.append({
                "data": datas[i].isoformat(),
                "rotulo": rotulo(datas[i]),
                "leitura": v,
                "consumo": consumo,
                "valor": round(consumo * TARIFA, 2) if consumo is not None else None,
                "flags": flags,
            })

        consumos = [p["consumo"] for p in serie if p["consumo"] is not None]
        # consumo do PERIODO ATUAL = consumo na ultima coluna de data
        ult_consumo = col_consumo[-1]
        sem_recente = reads[-1] is None
        # leitura mais recente disponivel (pode nao ser a ultima coluna)
        ultimo_lido = next((i for i in range(len(reads) - 1, -1, -1) if reads[i] is not None), None)
        ult_data = datas[ultimo_lido] if ultimo_lido is not None else None
        # variacao do periodo atual vs periodo anterior (colunas)
        var = None
        if col_consumo[-1] is not None and col_consumo[-2] is not None and col_consumo[-2] > 0:
            var = round((col_consumo[-1] - col_consumo[-2]) / col_consumo[-2] * 100, 1)

        areas.append({
            "nome": loc,
            "grupo": grupo_atual,
            "chave": grupo_atual + " | " + loc,
            "serie": serie,
            "resumo": {
                "consumo_total": sum(consumos) if consumos else 0,
                "valor_total": round(sum(consumos) * TARIFA, 2) if consumos else 0,
                "media": round(sum(consumos) / len(consumos), 1) if consumos else 0,
                "consumo_ultimo": ult_consumo,
                "valor_ultimo": round(ult_consumo * TARIFA, 2) if ult_consumo is not None else None,
                "variacao_pct": var,
                "leituras": len(serie),
                "ult_data": ult_data.isoformat() if ult_data else None,
                "ult_rotulo": rotulo(ult_data) if ult_data else None,
                "leitura_atual": reads[ultimo_lido] if ultimo_lido is not None else None,
                "sem_leitura_recente": sem_recente,
            },
        })

    # ---- serie agregada (evolucao do total) ----
    serie_total = []
    for i in range(len(col_nums)):
        if i == 0:
            continue  # 1a coluna nao tem consumo (sem anterior)
        serie_total.append({
            "data": datas[i].isoformat(),
            "rotulo": rotulo(datas[i]),
            "consumo": round(total_por_col[i]),
            "valor": round(total_por_col[i] * TARIFA, 2),
        })

    # ultimo periodo e o anterior (para KPIs)
    ult = serie_total[-1] if serie_total else None
    ant = serie_total[-2] if len(serie_total) >= 2 else None
    variacao_geral = None
    if ult and ant and ant["consumo"] > 0:
        variacao_geral = round((ult["consumo"] - ant["consumo"]) / ant["consumo"] * 100, 1)

    # ranking do ultimo periodo (consumo_ultimo por ponto, desc)
    ranking = sorted(
        [{"nome": a["nome"], "grupo": a["grupo"],
          "consumo": a["resumo"]["consumo_ultimo"] or 0,
          "valor": a["resumo"]["valor_ultimo"] or 0} for a in areas],
        key=lambda x: x["consumo"], reverse=True)
    maior = ranking[0] if ranking else None

    # resumo por grupo (ultimo periodo)
    grupos_ord = {g: {"nome": g, "qtd": 0, "consumo_ultimo": 0, "valor_ultimo": 0.0} for g in GRUPOS}
    for a in areas:
        g = grupos_ord.setdefault(a["grupo"], {"nome": a["grupo"], "qtd": 0, "consumo_ultimo": 0, "valor_ultimo": 0.0})
        g["qtd"] += 1
        cu = a["resumo"]["consumo_ultimo"]
        if cu is not None:
            g["consumo_ultimo"] += cu
            g["valor_ultimo"] += (a["resumo"]["valor_ultimo"] or 0)
    grupos_lista = [dict(v, valor_ultimo=round(v["valor_ultimo"], 2)) for v in grupos_ord.values() if v["qtd"] > 0]

    meta = wb.worksheets[0]  # 1a aba para pegar associacao
    assoc = meta.cell(1, 1).value

    saida = {
        "gerado_em": datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
        "fonte": os.path.relpath(arq, BASE_DIR).replace("\\", "/"),
        "condominio": {"associacao": assoc, "titulo": "Consumo de água — Áreas Comuns"},
        "tarifa": TARIFA,
        "periodo_inicial": rotulo(datas[0]) if datas else None,
        "periodo_final": rotulo(datas[-1]) if datas else None,
        "resumo_geral": {
            "consumo_ultimo": ult["consumo"] if ult else 0,
            "valor_ultimo": ult["valor"] if ult else 0,
            "rotulo_atual": ult["rotulo"] if ult else None,
            "rotulo_anterior": ant["rotulo"] if ant else None,
            "variacao_pct": variacao_geral,
            "n_pontos": len(areas),
            "n_grupos": len(grupos_lista),
            "tarifa": TARIFA,
            "maior": maior,
        },
        "serie_total": serie_total,
        "ranking": ranking,
        "grupos": grupos_lista,
        "areas": areas,
    }

    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    with open(SAIDA, "w", encoding="utf-8") as fp:
        fp.write("window.DADOS_AC = ")
        json.dump(saida, fp, ensure_ascii=False, separators=(",", ":"))
        fp.write(";\n")

    print(f"OK -> {os.path.relpath(SAIDA, BASE_DIR)}")
    print(f"  fonte: {saida['fonte']}")
    print(f"  periodo: {saida['periodo_inicial']} -> {saida['periodo_final']}  ({len(serie_total)+1} leituras)")
    print(f"  pontos de medicao: {len(areas)}  |  grupos: {len(grupos_lista)}  |  excluidos: {excluidos}")
    print(f"  consumo ultimo periodo ({saida['resumo_geral']['rotulo_atual']}): "
          f"{saida['resumo_geral']['consumo_ultimo']} m3  =  R$ {saida['resumo_geral']['valor_ultimo']}")
    print(f"  maior consumidor: {maior['nome']} ({maior['consumo']} m3)")
    print(f"  variacao vs anterior: {variacao_geral}%")
    print(f"  tamanho: {os.path.getsize(SAIDA)/1024:.0f} KB")


if __name__ == "__main__":
    main()
