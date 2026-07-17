# -*- coding: utf-8 -*-
"""
Converte os arquivos "PARA COBRANÇA" (.xlsx) para JSON.

- Procura recursivamente por arquivos cujo nome contenha "PARA COBRAN".
- Le a primeira planilha (tabela de cobranca por unidade).
- Mapeia as colunas pelo NOME do cabecalho (nao pela posicao), entao funciona
  tanto para planilhas com a coluna "Releitura" quanto para as que nao tem.
- Extrai metadados do topo (associacao, apuracao, datas) e a lista de unidades
  com leitura anterior/atual, consumo (m3), valor e dados do hidrometro.
- Gera um JSON por arquivo em ./json/ e um combinado (todos_para_cobranca.json).

Uso:
    python converter_para_json.py

Requer: openpyxl  (pip install openpyxl)
"""

import json
import re
import glob
import os

try:
    import openpyxl
except ImportError:
    raise SystemExit("openpyxl nao instalado. Rode: python -m pip install openpyxl")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(BASE_DIR, "json")

# Ordem canonica das chaves em cada registro de saida.
CHAVES_SAIDA = [
    "bloco",
    "unidade",
    "leitura_anterior",
    "leitura_atual",
    "releitura",
    "consumo_m3",
    "valor",
    "condicao_hidrometro",
    "estado_cavalete",
    "numero_lacre",
]


def limpar(valor):
    """Remove espacos extras de strings; mantem outros tipos como estao."""
    if isinstance(valor, str):
        v = valor.strip()
        return v if v != "" else None
    return valor


def classificar_cabecalho(h):
    """Mapeia o texto de um cabecalho para a chave de saida correspondente."""
    if not isinstance(h, str):
        return None
    t = h.strip().lower()
    if t == "bloco":
        return "bloco"
    if t == "unidade":
        return "unidade"
    if t == "anterior":
        return "leitura_anterior"
    if t.startswith("releitura"):
        return "releitura"
    if t in ("m³", "m3") or t.startswith("m³") or t.startswith("m3"):
        return "consumo_m3"
    if t == "valor":
        return "valor"
    if t.startswith("condi"):        # Condição do hidrômetro
        return "condicao_hidrometro"
    if t.startswith("estado"):       # Estado do cavalete
        return "estado_cavalete"
    if "lacre" in t:                 # Nº do lacre
        return "numero_lacre"
    return None


def achar_linha_cabecalho(ws, limite=30):
    """Retorna o numero da linha onde esta o cabecalho 'Bloco'."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=limite, values_only=True), 1):
        for celula in row:
            if isinstance(celula, str) and celula.strip() == "Bloco":
                return i
    return None


def mapear_colunas(ws, hdr):
    """Retorna {chave: numero_da_coluna} e o rotulo da coluna de leitura atual.

    A coluna de leitura atual tem cabecalho variavel (uma data, ex.: '28.05'),
    entao e identificada como a coluna imediatamente apos 'Anterior'.
    """
    mapa = {}
    for c in range(1, ws.max_column + 1):
        chave = classificar_cabecalho(ws.cell(hdr, c).value)
        if chave and chave not in mapa:
            mapa[chave] = c

    rotulo_atual = None
    if "leitura_anterior" in mapa:
        col_atual = mapa["leitura_anterior"] + 1
        # So assume como leitura_atual se aquela coluna ainda nao foi classificada.
        if col_atual not in mapa.values():
            mapa["leitura_atual"] = col_atual
            rot = ws.cell(hdr, col_atual).value
            if hasattr(rot, "strftime"):
                rot = rot.strftime("%d/%m/%Y")
            rotulo_atual = str(rot) if rot is not None else None
    return mapa, rotulo_atual


def extrair_data(texto):
    """Tenta extrair uma data (dd/mm/aaaa ou dd.mm.aaaa) de um texto."""
    if not isinstance(texto, str):
        return None
    m = re.search(r"(\d{2}[/.]\d{2}[/.]\d{4})", texto)
    return m.group(1).replace(".", "/") if m else None


def texto_celula(ws, linha, col=1):
    return limpar(ws.cell(linha, col).value)


def processar_arquivo(caminho):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb.worksheets[0]  # primeira planilha = tabela de cobranca

    hdr = achar_linha_cabecalho(ws)
    if hdr is None:
        raise ValueError(f"Cabecalho 'Bloco' nao encontrado em {caminho}")

    mapa, rotulo_atual = mapear_colunas(ws, hdr)

    col_releitura = mapa.get("releitura")
    rotulo_releitura = limpar(ws.cell(hdr, col_releitura).value) if col_releitura else None

    # Metadados do topo da planilha.
    linha4 = texto_celula(ws, 4)  # "L.Anterior = Leitura feita em ..."
    linha5 = texto_celula(ws, 5)  # "L.Atual = Leitura feita em ..."

    rel = os.path.relpath(caminho, BASE_DIR).replace("\\", "/")
    diretorio = os.path.dirname(rel)

    meta = {
        "arquivo_origem": rel,
        "pasta": diretorio.split("/")[0] if diretorio else "(raiz)",
        "planilha": ws.title,
        "associacao": texto_celula(ws, 1),
        "titulo": texto_celula(ws, 2),
        "apuracao": texto_celula(ws, 3),
        "descricao_leitura_anterior": linha4,
        "descricao_leitura_atual": linha5,
        "data_leitura_anterior": extrair_data(linha4),
        "data_leitura_atual": extrair_data(linha5) or (rotulo_atual if "/" in str(rotulo_atual) else None),
        "rotulo_coluna_atual": rotulo_atual,
        "rotulo_coluna_releitura": rotulo_releitura,
        "possui_releitura": col_releitura is not None,
    }

    registros = []
    for r in range(hdr + 1, ws.max_row + 1):
        bloco = limpar(ws.cell(r, mapa["bloco"]).value) if "bloco" in mapa else None
        unidade = limpar(ws.cell(r, mapa["unidade"]).value) if "unidade" in mapa else None
        # Linha de dados valida = tem Bloco e Unidade (exclui vazias e "total").
        if bloco is None or unidade is None:
            continue
        registro = {}
        for chave in CHAVES_SAIDA:
            col = mapa.get(chave)
            registro[chave] = limpar(ws.cell(r, col).value) if col else None
        registros.append(registro)

    meta["total_registros"] = len(registros)
    return {"metadados": meta, "leituras": registros}


def nome_saida(caminho):
    """Gera um nome de arquivo unico a partir da pasta (ou do nome, se na raiz)."""
    rel = os.path.relpath(caminho, BASE_DIR)
    partes = rel.split(os.sep)
    base = partes[0] if len(partes) > 1 else os.path.splitext(partes[0])[0]
    slug = re.sub(r"[^0-9A-Za-zÀ-ÿ]+", "_", base).strip("_")
    return f"{slug}.json"


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    padrao = os.path.join(BASE_DIR, "**", "*PARA COBRAN*.xlsx")
    arquivos = sorted(glob.glob(padrao, recursive=True))

    if not arquivos:
        print("Nenhum arquivo 'PARA COBRANCA' encontrado.")
        return

    combinado = []
    usados = {}
    for caminho in arquivos:
        dados = processar_arquivo(caminho)
        nome = nome_saida(caminho)
        # Evita sobrescrever caso dois caminhos gerem o mesmo nome.
        if nome in usados:
            usados[nome] += 1
            raiz, ext = os.path.splitext(nome)
            nome = f"{raiz}_{usados[nome]}{ext}"
        else:
            usados[nome] = 1
        saida = os.path.join(OUT_DIR, nome)
        with open(saida, "w", encoding="utf-8") as fp:
            json.dump(dados, fp, ensure_ascii=False, indent=2)
        combinado.append(dados)
        print(f"OK  {dados['metadados']['arquivo_origem']}")
        print(f"    -> json/{nome}  ({dados['metadados']['total_registros']} unidades, "
              f"releitura={'sim' if dados['metadados']['possui_releitura'] else 'nao'})")

    combinado_path = os.path.join(OUT_DIR, "todos_para_cobranca.json")
    with open(combinado_path, "w", encoding="utf-8") as fp:
        json.dump(combinado, fp, ensure_ascii=False, indent=2)
    print(f"\nCombinado: json/todos_para_cobranca.json  ({len(combinado)} arquivos)")


if __name__ == "__main__":
    main()
